"""Excel 파일 직접 읽기/쓰기 도구 (Excel 애플리케이션 없이).

fs 도구를 활용하여 파일 I/O를 수행하고, openpyxl을 사용하여 Excel 파일을 처리합니다.
모든 파일 접근은 fs 도구의 보안 정책을 따릅니다.
"""

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path

from django.conf import settings
from pydantic import Field

from pyhub.mcptools import mcp
from pyhub.mcptools.fs import core as fs_core


def _get_enabled_excel_tools():
    """Lazy evaluation of Excel tools enablement."""
    # Docker 환경에서 Excel 도구 비활성화
    return settings.USE_EXCEL_TOOLS and settings.FS_LOCAL_HOME is not None


# Excel 파일 확장자
EXCEL_FILE_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb"}


def is_excel_file(file_path: str) -> bool:
    """파일이 Excel 파일인지 확인."""
    return Path(file_path).suffix.lower() in EXCEL_FILE_EXTENSIONS


def csv_loads(data):
    """CSV 문자열을 파싱."""
    if not data:
        return []
    reader = csv.reader(StringIO(data))
    return list(reader)


def json_dumps(obj):
    """객체를 JSON 문자열로 변환."""
    return json.dumps(obj, ensure_ascii=False, indent=2)


@mcp.tool(enabled=lambda: _get_enabled_excel_tools())
async def file__excel_read(
    file_path: str = Field(description="Absolute path to Excel file"),
    sheet_name: str = Field(default="", description="Sheet name to read (first sheet if empty)"),
    range: str = Field(default="", description="Cell range to read (e.g., A1:C10, all cells if empty)"),
    data_only: bool = Field(default=True, description="True: formula results, False: formula strings"),
) -> str:
    """Read Excel file directly via file I/O without Excel application.

    This tool uses openpyxl library to read Excel files directly.
    All file access follows fs tool security policies.
    Use this for fast bulk read operations when Excel application is not available or needed.

    Returns:
        Data in CSV format
    """
    # 파일 확장자 확인
    if not is_excel_file(file_path):
        raise ValueError(f"지원하지 않는 파일 형식입니다: {Path(file_path).suffix}")

    try:
        # fs core로 파일 읽기 (바이너리)
        excel_bytes = await fs_core.read_file_binary(file_path)

        # openpyxl은 별도 설치가 필요하므로 동적 임포트
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "openpyxl이 설치되어 있지 않습니다. 'pip install pyhub-mcptools[excel]' 명령으로 설치해주세요."
            ) from e

        # Excel 파일 로드
        wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=data_only)

        # 시트 선택
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 범위 파싱
        if range:
            # 간단한 범위 파싱 (예: A1:C10)
            if ":" in range:
                start, end = range.split(":")
                min_row, min_col = ws[start].row, ws[start].column
                max_row, max_col = ws[end].row, ws[end].column
            else:
                # 단일 셀
                cell = ws[range]
                min_row = max_row = cell.row
                min_col = max_col = cell.column
        else:
            # 전체 범위
            min_row = ws.min_row
            max_row = ws.max_row
            min_col = ws.min_column
            max_col = ws.max_column

        # 데이터 읽기
        data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            data.append(list(row))

        # CSV로 변환
        output = StringIO()
        writer = csv.writer(output)
        writer.writerows(data)

        return output.getvalue()

    except Exception as e:
        raise ValueError(f"Excel 파일 읽기 실패: {str(e)}") from e


@mcp.tool(enabled=lambda: _get_enabled_excel_tools())
async def file__excel_write(
    file_path: str = Field(description="Absolute path to save Excel file"),
    data: str = Field(description="Data in CSV format"),
    sheet_name: str = Field(default="Sheet1", description="Sheet name"),
    overwrite: bool = Field(default=False, description="Whether to overwrite existing file"),
    header_format: bool = Field(
        default=False, description="Apply header formatting to first row (bold, background color)"
    ),
) -> str:
    """Write CSV data to Excel file directly via file I/O without Excel application.

    This tool uses openpyxl library to create Excel files directly.
    All file access follows fs tool security policies.
    Use this for fast bulk write operations when Excel application is not available or needed.

    Returns:
        Success message with file path
    """
    # 파일 확장자 확인
    if not is_excel_file(file_path):
        raise ValueError(f"지원하지 않는 파일 형식입니다: {Path(file_path).suffix}")

    try:
        # openpyxl 임포트
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise ImportError(
                "openpyxl이 설치되어 있지 않습니다. 'pip install pyhub-mcptools[excel]' 명령으로 설치해주세요."
            ) from e

        # CSV 데이터 파싱
        rows = list(csv_loads(data))
        if not rows:
            raise ValueError("데이터가 비어있습니다")

        # 새 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 데이터 쓰기
        for row in rows:
            ws.append(row)

        # 헤더 서식 적용 (옵션)
        if header_format and rows:
            try:
                from openpyxl.styles import Alignment, Font, PatternFill

                # 헤더 스타일 정의
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                # 첫 번째 행에 서식 적용
                for col_idx in range(1, len(rows[0]) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 열 너비 자동 조정
                for col_idx, col_data in enumerate(rows[0], 1):
                    max_length = len(str(col_data))
                    for row in rows[1:]:
                        if col_idx <= len(row):
                            max_length = max(max_length, len(str(row[col_idx - 1])))
                    adjusted_width = min(max_length + 2, 50)  # 최대 너비 50
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
            except ImportError:
                # 스타일 모듈이 없으면 무시
                pass

        # 메모리에 저장
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # fs core로 파일 쓰기
        # 파일 존재 여부 확인 (overwrite가 False인 경우)
        if not overwrite:
            try:
                # 파일이 이미 존재하는지 확인
                await fs_core.read_file_binary(file_path)
                raise ValueError(f"파일이 이미 존재합니다: {file_path}")
            except ValueError as e:
                if "not found" not in str(e).lower() and "no such file" not in str(e).lower():
                    raise

        excel_content = excel_buffer.getvalue()
        await fs_core.write_file(file_path, excel_content)

        return f"Excel 파일이 저장되었습니다: {file_path}"

    except Exception as e:
        raise ValueError(f"Excel 파일 쓰기 실패: {str(e)}") from e


@mcp.tool(enabled=lambda: _get_enabled_excel_tools())
async def file__excel_info(
    file_path: str = Field(description="Absolute path to Excel file"),
) -> str:
    """Get Excel file information directly via file I/O without Excel application.

    Returns sheet list, dimensions of each sheet, and other file metadata.
    Use this for quick file inspection when Excel application is not available or needed.

    Returns:
        File information in JSON format
    """
    # 파일 확장자 확인
    if not is_excel_file(file_path):
        raise ValueError(f"지원하지 않는 파일 형식입니다: {Path(file_path).suffix}")

    try:
        # fs core로 파일 읽기
        excel_bytes = await fs_core.read_file_binary(file_path)

        # openpyxl 임포트
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "openpyxl이 설치되어 있지 않습니다. 'pip install pyhub-mcptools[excel]' 명령으로 설치해주세요."
            ) from e

        # Excel 파일 로드
        wb = load_workbook(BytesIO(excel_bytes), read_only=True)

        # 파일 정보 수집
        info = {
            "file_path": file_path,
            "file_name": Path(file_path).name,
            "file_size": len(excel_bytes),
            "sheet_count": len(wb.sheetnames),
            "sheets": [],
        }

        # 각 시트 정보
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "cells": ws.max_row * ws.max_column if ws.max_row and ws.max_column else 0,
            }
            info["sheets"].append(sheet_info)

        return json_dumps(info)

    except Exception as e:
        raise ValueError(f"Excel 파일 정보 조회 실패: {str(e)}") from e


@mcp.tool(enabled=lambda: _get_enabled_excel_tools())
async def file__excel_convert(
    source_path: str = Field(description="Absolute path to source Excel file"),
    target_path: str = Field(description="Absolute path to save converted file"),
    target_format: str = Field(
        default="",
        description="Target format (auto-detect from extension if empty). Supported: xlsx, xls",
    ),
) -> str:
    """Convert Excel file between formats directly via file I/O without Excel application.

    Primarily used for .xls (legacy) ↔ .xlsx (modern) conversion.
    Use this for fast format conversion when Excel application is not available or needed.

    Returns:
        Success message with file path
    """
    # 파일 확장자 확인
    if not is_excel_file(source_path):
        raise ValueError(f"지원하지 않는 원본 파일 형식입니다: {Path(source_path).suffix}")

    # 대상 형식 결정
    if not target_format:
        target_suffix = Path(target_path).suffix.lower()
        if target_suffix == ".xlsx":
            target_format = "xlsx"
        elif target_suffix == ".xls":
            target_format = "xls"
        else:
            raise ValueError(f"지원하지 않는 대상 파일 형식입니다: {target_suffix}")

    if target_format not in ["xlsx", "xls"]:
        raise ValueError(f"지원하지 않는 변환 형식입니다: {target_format}")

    try:
        # fs core로 파일 읽기
        excel_bytes = await fs_core.read_file_binary(source_path)

        # 필요한 라이브러리 임포트
        try:
            from openpyxl import Workbook, load_workbook

            if target_format == "xls" or Path(source_path).suffix.lower() == ".xls":
                import xlrd
                import xlwt
        except ImportError as e:
            raise ImportError(
                "필요한 라이브러리가 설치되어 있지 않습니다. "
                "'pip install pyhub-mcptools[excel] xlrd xlwt' 명령으로 설치해주세요."
            ) from e

        # 변환 처리
        if Path(source_path).suffix.lower() == ".xls" and target_format == "xlsx":
            # xls → xlsx 변환
            import xlrd

            # xls 파일 읽기
            book = xlrd.open_workbook(file_contents=excel_bytes)

            # 새 xlsx 워크북 생성
            new_wb = Workbook()
            new_wb.remove(new_wb.active)  # 기본 시트 제거

            for sheet_name in book.sheet_names():
                sheet = book.sheet_by_name(sheet_name)
                new_ws = new_wb.create_sheet(title=sheet_name)

                # 데이터 복사
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        value = sheet.cell_value(row, col)
                        new_ws.cell(row=row + 1, column=col + 1, value=value)

            # 메모리에 저장
            output_buffer = BytesIO()
            new_wb.save(output_buffer)
            output_buffer.seek(0)

        elif Path(source_path).suffix.lower() == ".xlsx" and target_format == "xls":
            # xlsx → xls 변환
            import xlwt

            # xlsx 파일 읽기
            wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)

            # 새 xls 워크북 생성
            new_book = xlwt.Workbook()

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                new_sheet = new_book.add_sheet(sheet_name)

                # 데이터 복사 (xls는 65536행, 256열 제한)
                max_row = min(ws.max_row, 65536)
                max_col = min(ws.max_column, 256)

                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
                ):
                    for col_idx, value in enumerate(row):
                        if value is not None:
                            new_sheet.write(row_idx, col_idx, value)

            # 메모리에 저장
            output_buffer = BytesIO()
            new_book.save(output_buffer)
            output_buffer.seek(0)

        else:
            # 같은 형식으로 복사 (재저장)
            if Path(source_path).suffix.lower() == ".xlsx":
                wb = load_workbook(BytesIO(excel_bytes))
                output_buffer = BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
            else:
                raise ValueError("같은 형식으로의 변환은 현재 지원하지 않습니다.")

        # fs core로 파일 쓰기
        await fs_core.write_file(target_path, output_buffer.getvalue())

        return f"Excel 파일이 변환되었습니다: {source_path} → {target_path}"

    except Exception as e:
        raise ValueError(f"Excel 파일 변환 실패: {str(e)}") from e


@mcp.tool(enabled=lambda: _get_enabled_excel_tools())
async def file__excel_merge(
    file_paths: str = Field(description="Excel file paths to merge (comma-separated)"),
    target_path: str = Field(description="Absolute path to save merged file"),
    merge_mode: str = Field(
        default="sheets",
        description="Merge mode: 'sheets' (each file as separate sheet) or 'append' (append data vertically)",
    ),
    sheet_prefix: str = Field(
        default="",
        description="Sheet name prefix (only used in sheets mode)",
    ),
) -> str:
    """Merge multiple Excel files directly via file I/O without Excel application.

    Supports two merge modes:
    - sheets: Each file becomes a separate sheet
    - append: All data appended to single sheet
    Use this for efficient bulk file merging when Excel application is not available or needed.

    Returns:
        Success message with merged file information
    """
    # 파일 경로 파싱
    paths = [p.strip() for p in file_paths.split(",") if p.strip()]
    if len(paths) < 2:
        raise ValueError("병합하려면 최소 2개 이상의 파일이 필요합니다.")

    # 모든 파일이 Excel 파일인지 확인
    for path in paths:
        if not is_excel_file(path):
            raise ValueError(f"지원하지 않는 파일 형식입니다: {path}")

    if merge_mode not in ["sheets", "append"]:
        raise ValueError(f"지원하지 않는 병합 모드입니다: {merge_mode}")

    try:
        # openpyxl 임포트
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as e:
            raise ImportError(
                "openpyxl이 설치되어 있지 않습니다. 'pip install pyhub-mcptools[excel]' 명령으로 설치해주세요."
            ) from e

        # 새 워크북 생성
        merged_wb = Workbook()

        if merge_mode == "sheets":
            # sheets 모드: 각 파일을 별도 시트로
            merged_wb.remove(merged_wb.active)  # 기본 시트 제거

            for file_path in paths:
                # fs core로 파일 읽기
                excel_bytes = await fs_core.read_file_binary(file_path)
                wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)

                # 각 시트를 복사
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    # 새 시트 이름 생성
                    base_name = Path(file_path).stem
                    if sheet_prefix:
                        new_sheet_name = f"{sheet_prefix}_{base_name}_{sheet_name}"
                    else:
                        new_sheet_name = f"{base_name}_{sheet_name}"

                    # 시트 이름이 중복되지 않도록 조정
                    final_name = new_sheet_name
                    counter = 1
                    while final_name in merged_wb.sheetnames:
                        final_name = f"{new_sheet_name}_{counter}"
                        counter += 1

                    # 새 시트 생성 및 데이터 복사
                    new_ws = merged_wb.create_sheet(title=final_name[:31])  # Excel 시트 이름은 31자 제한

                    for row in ws.iter_rows(values_only=True):
                        new_ws.append(list(row))

        else:  # append 모드
            # append 모드: 모든 데이터를 하나의 시트에
            merged_ws = merged_wb.active
            merged_ws.title = "Merged"

            first_file = True

            for file_path in paths:
                # fs core로 파일 읽기
                excel_bytes = await fs_core.read_file_binary(file_path)
                wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)

                # 첫 번째 시트의 데이터만 사용
                ws = wb.active

                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    # 첫 번째 파일의 헤더는 포함, 나머지 파일의 헤더는 건너뛰기
                    if row_idx == 0 and not first_file:
                        continue

                    merged_ws.append(list(row))

                first_file = False

        # 메모리에 저장
        output_buffer = BytesIO()
        merged_wb.save(output_buffer)
        output_buffer.seek(0)

        # fs core로 파일 쓰기
        await fs_core.write_file(target_path, output_buffer.getvalue())

        # 결과 정보 생성
        result_info = {
            "merged_file": target_path,
            "source_files": len(paths),
            "merge_mode": merge_mode,
            "sheet_count": len(merged_wb.sheetnames),
        }

        return f"Excel 파일 병합 완료: {json_dumps(result_info)}"

    except Exception as e:
        raise ValueError(f"Excel 파일 병합 실패: {str(e)}") from e


@mcp.tool(enabled=lambda: _get_enabled_excel_tools())
async def file__excel_format(
    file_path: str = Field(description="Absolute path to Excel file to format"),
    sheet_name: str = Field(default="", description="Sheet name to format (first sheet if empty)"),
    format_options: str = Field(
        description="Format options in JSON. Example: "
        '{"header_row": true, "auto_filter": true, "freeze_panes": "A2", '
        '"column_widths": {"A": 20, "B": 15}, "number_formats": {"C": "#,##0.00"}}'
    ),
    output_path: str = Field(default="", description="Save path (overwrite original if empty)"),
) -> str:
    """Apply formatting to Excel file directly via file I/O without Excel application.

    Supported format options:
    - header_row: Format first row as header
    - auto_filter: Apply auto filter
    - freeze_panes: Freeze panes (e.g., "A2" freezes first row)
    - column_widths: Set column widths (e.g., {"A": 20, "B": 15})
    - number_formats: Set number formats (e.g., {"C": "#,##0.00"})
    Use this for batch formatting operations when Excel application is not available or needed.

    Returns:
        Success message with file path
    """
    # 파일 확장자 확인
    if not is_excel_file(file_path):
        raise ValueError(f"지원하지 않는 파일 형식입니다: {Path(file_path).suffix}")

    try:
        # fs core로 파일 읽기
        excel_bytes = await fs_core.read_file_binary(file_path)

        # openpyxl 임포트
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as e:
            raise ImportError(
                "openpyxl이 설치되어 있지 않습니다. 'pip install pyhub-mcptools[excel]' 명령으로 설치해주세요."
            ) from e

        # 서식 옵션 파싱
        import json

        try:
            options = json.loads(format_options)
        except json.JSONDecodeError as e:
            raise ValueError(f"잘못된 JSON 형식입니다: {str(e)}") from e

        # Excel 파일 로드
        wb = load_workbook(BytesIO(excel_bytes))

        # 시트 선택
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 헤더 행 서식
        if options.get("header_row", False):
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # 첫 번째 행
                if cell.value is not None:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        # 자동 필터
        if options.get("auto_filter", False):
            ws.auto_filter.ref = ws.dimensions

        # 틀 고정
        if "freeze_panes" in options:
            ws.freeze_panes = options["freeze_panes"]

        # 열 너비 설정
        if "column_widths" in options:
            for col_letter, width in options["column_widths"].items():
                ws.column_dimensions[col_letter].width = width

        # 숫자 형식 설정
        if "number_formats" in options:
            for col_letter, format_code in options["number_formats"].items():
                col_idx = ord(col_letter) - ord("A") + 1
                for row in range(2, ws.max_row + 1):  # 헤더 제외
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = format_code

        # 메모리에 저장
        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        # fs core로 파일 쓰기
        save_path = output_path if output_path else file_path
        await fs_core.write_file(save_path, output_buffer.getvalue())

        return f"Excel 파일에 서식이 적용되었습니다: {save_path}"

    except Exception as e:
        raise ValueError(f"Excel 파일 서식 적용 실패: {str(e)}") from e
